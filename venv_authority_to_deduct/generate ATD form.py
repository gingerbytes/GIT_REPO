from tkinter import *
from tkinter.filedialog import askopenfilename
from openpyxl import load_workbook
from datetime import datetime
import fitz
from os import startfile

'''
# generates dictionary of letter:number
import string
dict(zip(string.ascii_lowercase, range(1,27)))
'''

class Cell():
    def __init__(self, value='', row=0, col=0, coor=[]):
        self.value = value
        self.row = row
        self.col = col
        self.coor = []

def find_cell(value, wsheet):
    for row in wsheet.rows:
        for col in range(0, wsheet.max_column):
            if row[col].value == value:
                return(row[col].row, col + 1)
    else:
        print(f'{value} not found in the {ws_name} worksheet.')

def emp_lookup(emp_id):
    fname.row, fname.col = find_cell('FirstName', ws)
    mname.row, mname.col = find_cell('MiddleName', ws)
    lname.row, lname.col = find_cell('Last Name', ws)
    depno.row, depno.col = find_cell('Local Department', ws)
    ccenter.row, ccenter.col = find_cell('SmSCostCenter', ws)
    years.row, years.col = find_cell('AdjDateofHire', ws)
    empid.row, empid.col = find_cell(int(emp_id), ws)

    now = datetime.today()
    fname.value = ws.cell(row = empid.row, column = fname.col).value
    mname.value = ws.cell(row = empid.row, column = mname.col).value
    lname.value = ws.cell(row = empid.row, column = lname.col).value
    years.value = str(ws.cell(row = empid.row, column = years.col).value)

    # fields needed in the form
    name.value = (f'{fname.value} {mname.value} {lname.value }')
    depno.value = ws.cell(row = empid.row, column = depno.col).value
    empid.value = emp_id
    years.value = int(now.strftime("%Y")) - int(years.value[:4])

    print(f'{name.value } from {depno.value } with Emp ID {empid.value} has been with Infor for {years.value } years.')

def pdf_add_emp_info(fname, newname):

    name = 'Gerard Dangca'
    depno = 'SaaS'
    empid = '29849'
    years = 9

    doc = fitz.open(fname)
    page = doc[0]

    where = fitz.Point(50, 100, ) 
    text = name

    page._cleanContents()
    page.insertText(where, text, rotate=0, render_mode=2, overlay = True)

    '''
    page.insertText(where, text,
                    fontname=fname,    # arbitrary if fontfile given
                    fontfile=ffile,    # any file containing a font
                    fontsize=11,       # default
                    rotate=0,          # rotate text
                    color=(0, 0, 1),   # some color (blue)
                    overlay=True)      # text in foreground
    '''

    doc.save(newname)

fname = Cell() # do not shortcut class() declaration.
mname = Cell() # ex. fname, mname = (Class(),)*2
lname = Cell() # this causes problems on values of class attrib. 
depno = Cell()
years = Cell()
ccenter = Cell()
empid = Cell()
name = Cell()

ws_name = 'Details'  # this worksheet contains the employee details.
tk = Tk()
tk.withdraw()  # hides tk dialog box.
#EE_list = askopenfilename(title="Open EE List Excel File.", filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))
#EE_list = 'C:/Users/gdangca/OneDrive - Infor/Projects/Project - Python/GIT_REPO/venv_authority_to_deduct/Files/EE List 06-19-2020 nosal.xlsx'
EE_list = 'D:/DOCUMENTS/GERARD/Python - Poject/GIT_REPO/venv_authority_to_deduct/Files/EE List 06-19-2020 nosal.xlsx'

wb = load_workbook(EE_list, read_only=True)
for i in range(len(wb.sheetnames)):
    if wb.sheetnames[i] == ws_name:
        wb.active = i
        ws = wb.active
        break
else:
    print(f'{ws_name} worksheet is not in the excel file. Try another file')

eid = '30011'

'''
while True:
    eid = input("Please input employee ID of Applicant: ")
    if eid.isdigit():
        break
'''

#emp_lookup(eid)

pdf_atd= askopenfilename(title="Open PDF Authority to Deduct Form.", filetypes=(("PDF files", "*.pdf"), ("all files", "*.*")))
pdf_atd_eid = pdf_atd[:len(pdf_atd) - 4] + '_' + eid + '.pdf'
pdf_add_emp_info(pdf_atd, pdf_atd_eid)

startfile(pdf_atd_eid)  # open the signed PDF.