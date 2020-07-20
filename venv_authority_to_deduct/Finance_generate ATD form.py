from tkinter import *
from tkinter.filedialog import askopenfilename
from openpyxl import load_workbook
from datetime import datetime
import fitz
from os import startfile

'''
# generates dictionary of letter:number
import string
dict(zip(string.ascii_lowercase, range(1,27)))
'''

class Cell():
    def __init__(self, value='', row=0, col=0):
        self.value = value
        self.row = row
        self.col = col

def find_cell(value, wsheet):
    for row in wsheet.rows:
        for col in range(0, wsheet.max_column):
            if row[col].value == value:
                return(row[col].row, col + 1)
    else:
        print(f'{value} not found in the {ws_name} worksheet.')

def emp_lookup(emp_id):
    fname.row, fname.col = find_cell('FirstName', ws)
    mname.row, mname.col = find_cell('MiddleName', ws)
    lname.row, lname.col = find_cell('Last Name', ws)
    depno.row, depno.col = find_cell('Local Department', ws)
    ccenter.row, ccenter.col = find_cell('SmSCostCenter', ws)
    years.row, years.col = find_cell('AdjDateofHire', ws)
    empid.row, empid.col = find_cell(int(emp_id), ws)

    now = datetime.today()
    fname.value = ws.cell(row = empid.row, column = fname.col).value
    mname.value = ws.cell(row = empid.row, column = mname.col).value
    lname.value = ws.cell(row = empid.row, column = lname.col).value
    years.value = str(ws.cell(row = empid.row, column = years.col).value)

    # fields needed in the form
    name.value = (f'{fname.value} {mname.value} {lname.value }')
    depno.value = ws.cell(row = empid.row, column = depno.col).value
    empid.value = emp_id
    years.value = int(now.strftime("%Y")) - int(years.value[:4])

    print(f'{name.value } from {depno.value } with Emp ID {empid.value} has been with Infor for {years.value } years.')
    return

def pdf_add_emp_info(name='NAME', depno='DEPARTMENT', empid='1234', years='10' ):

    appp_xy = fitz.Point(113, 81) # Applicant
    depno_xy = fitz.Point(113, 93) # Dept No.
    empid_xy = fitz.Point(363, 81) # Employee ID No.
    year_xy = fitz.Point(363, 93) # Length of Service

    fname = askopenfilename(title="Open PDF Authority to Deduct Form.", filetypes=(("PDF files", "*.pdf"), ("all files", "*.*")))
    fname_empid = fname[:len(fname) - 4] + '_' + empid + '.pdf'
    doc = fitz.open(fname)
    page = doc[0]
    page.insertText(appp_xy, name, rotate=0, fontsize=9, render_mode=0, overlay = True)
    page.insertText(depno_xy, depno, rotate=0, fontsize=9, render_mode=0, overlay = True)
    page.insertText(empid_xy, empid, rotate=0, fontsize=9, render_mode=0, overlay = True)
    page.insertText(year_xy, years, rotate=0, fontsize=9, render_mode=0, overlay = True)

    doc.save(fname_empid)
    startfile(fname_empid)  # open the signed PDF.
    return

fname = Cell() # do not shortcut class() declaration.
mname = Cell() # ex. fname, mname = (Class(),)*2
lname = Cell() # this causes problems on values of class attrib. 
depno = Cell()
years = Cell()
ccenter = Cell()
empid = Cell()
name = Cell()

ws_name = 'Details'  # this worksheet contains the employee details.
tk = Tk()
tk.withdraw()  # hides tk dialog box.
EE_list = askopenfilename(title="Open EE List Excel File.", filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))
#EE_list = 'C:/Users/gdangca/OneDrive - Infor/Projects/Project - Python/GIT_REPO/venv_authority_to_deduct/Files/EE List 06-19-2020 nosal.xlsx'
#EE_list = 'D:/DOCUMENTS/GERARD/Python - Poject/GIT_REPO/venv_authority_to_deduct/Files/EE List 06-19-2020 nosal.xlsx'

wb = load_workbook(EE_list, read_only=True)
for i in range(len(wb.sheetnames)):
    if wb.sheetnames[i] == ws_name:
        wb.active = i
        ws = wb.active
        break
else:
    print(f'{ws_name} worksheet is not in the excel file. Try another file')

eid = 'abc'
while not eid.isdigit():
    eid = input("Please input employee ID of Applicant: ")

emp_lookup(eid)
pdf_add_emp_info(name.value, depno.value, empid.value, str(years.value))