from tkinter import *
from tkinter.filedialog import askdirectory
import openpyxl
import xlrd

summary_gerard = "Summary of Cash Claims - Gerard.xlsx"
summary_dennis = "Summary of Cash Claims - Dennis.xlsx"
#summary_gerard = "test.xlsx"
summary_gerard = "Summary of Cash Claims - Gerard2.xls"

tk = Tk()
tk.withdraw()  # hides tk dialog box.
#path = askdirectory() # ask where the reports are located.
path = "D:/DOCUMENTS/GERARD/Python - Poject/GIT_REPO/venv_cash_claim/Reports"
filepath = path + '/' + summary_gerard

wb = openpyxl.Workbook()
ws = wb.active

#wb = openpyxl.load_workbook(filepath) 
#ws = wb.active
#cell_obj = ws.cell(row = 2, column = 11) 
#print(cell_obj.value)

#XLRD
book = xlrd.open_workbook(filepath)
first_sheet = book.sheet_by_index(0)
cell = first_sheet.cell(2,11)
print(cell)
print(cell.values)